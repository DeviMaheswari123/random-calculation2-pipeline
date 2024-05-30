# randomCalc.py

from openpyxl import Workbook
from random import randint

def main():
    # Generate random numbers
    num1 = randint(1, 100)
    num2 = randint(1, 100)

    # Perform the calculation
    result = num1 + num2

    # Create a new workbook and select the active worksheet
    wb = Workbook()
    ws = wb.active

    # Write the numbers and result to the worksheet
    ws['A1'] = num1
    ws['B1'] = '+'
    ws['C1'] = num2
    ws['D1'] = '='
    ws['E1'] = result

    # Save the workbook
    wb.save('random_calculation3.xlsx')

if __name__ == "__main__":
    main()
